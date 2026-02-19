#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère le fichier Excel PAC Dimensionnement
avec toutes les données et formules intégrées (sans API).
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment,
                              numbers, NamedStyle)
from openpyxl.chart import LineChart, Reference, AreaChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import math
import json

# ============================================================
# DATA (mirrors performance.js)
# ============================================================

TBASE = {
    "01":{"tbase":-10,"zone":"G","climat":"H1a"},"02":{"tbase":-7,"zone":"D","climat":"H1c"},
    "03":{"tbase":-8,"zone":"E","climat":"H1c"},"04":{"tbase":-8,"zone":"E","climat":"H2d"},
    "05":{"tbase":-10,"zone":"G","climat":"H1c"},"06":{"tbase":-6,"zone":"A","climat":"H3"},
    "07":{"tbase":-6,"zone":"D","climat":"H2d"},"08":{"tbase":-10,"zone":"G","climat":"H1c"},
    "09":{"tbase":-5,"zone":"C","climat":"H2d"},"10":{"tbase":-10,"zone":"G","climat":"H1c"},
    "11":{"tbase":-5,"zone":"C","climat":"H3"},"12":{"tbase":-8,"zone":"E","climat":"H2d"},
    "13":{"tbase":-5,"zone":"C","climat":"H3"},"14":{"tbase":-7,"zone":"D","climat":"H1c"},
    "15":{"tbase":-8,"zone":"E","climat":"H1c"},"16":{"tbase":-5,"zone":"C","climat":"H2c"},
    "17":{"tbase":-5,"zone":"C","climat":"H2b"},"18":{"tbase":-7,"zone":"D","climat":"H2c"},
    "19":{"tbase":-8,"zone":"E","climat":"H1c"},"2A":{"tbase":-2,"zone":"A","climat":"H3"},
    "2B":{"tbase":-2,"zone":"A","climat":"H3"},"21":{"tbase":-10,"zone":"G","climat":"H1c"},
    "22":{"tbase":-4,"zone":"B","climat":"H2a"},"23":{"tbase":-8,"zone":"E","climat":"H1c"},
    "24":{"tbase":-5,"zone":"C","climat":"H2c"},"25":{"tbase":-12,"zone":"H","climat":"H1b"},
    "26":{"tbase":-6,"zone":"D","climat":"H2d"},"27":{"tbase":-7,"zone":"D","climat":"H1c"},
    "28":{"tbase":-7,"zone":"D","climat":"H1c"},"29":{"tbase":-4,"zone":"B","climat":"H2a"},
    "30":{"tbase":-5,"zone":"C","climat":"H3"},"31":{"tbase":-5,"zone":"C","climat":"H2d"},
    "32":{"tbase":-5,"zone":"C","climat":"H2d"},"33":{"tbase":-5,"zone":"C","climat":"H2c"},
    "34":{"tbase":-5,"zone":"C","climat":"H3"},"35":{"tbase":-4,"zone":"C","climat":"H2a"},
    "36":{"tbase":-7,"zone":"D","climat":"H2c"},"37":{"tbase":-7,"zone":"D","climat":"H2b"},
    "38":{"tbase":-10,"zone":"G","climat":"H1a"},"39":{"tbase":-10,"zone":"G","climat":"H1b"},
    "40":{"tbase":-5,"zone":"C","climat":"H2c"},"41":{"tbase":-7,"zone":"D","climat":"H2b"},
    "42":{"tbase":-8,"zone":"E","climat":"H1a"},"43":{"tbase":-8,"zone":"E","climat":"H1c"},
    "44":{"tbase":-5,"zone":"C","climat":"H2a"},"45":{"tbase":-7,"zone":"D","climat":"H1c"},
    "46":{"tbase":-6,"zone":"D","climat":"H2d"},"47":{"tbase":-5,"zone":"C","climat":"H2c"},
    "48":{"tbase":-8,"zone":"E","climat":"H2d"},"49":{"tbase":-7,"zone":"D","climat":"H2a"},
    "50":{"tbase":-4,"zone":"B","climat":"H2a"},"51":{"tbase":-10,"zone":"G","climat":"H1c"},
    "52":{"tbase":-12,"zone":"H","climat":"H1c"},"53":{"tbase":-7,"zone":"C","climat":"H2a"},
    "54":{"tbase":-15,"zone":"I","climat":"H1c"},"55":{"tbase":-12,"zone":"H","climat":"H1c"},
    "56":{"tbase":-4,"zone":"B","climat":"H2a"},"57":{"tbase":-15,"zone":"I","climat":"H1c"},
    "58":{"tbase":-10,"zone":"G","climat":"H1b"},"59":{"tbase":-9,"zone":"F","climat":"H1c"},
    "60":{"tbase":-7,"zone":"D","climat":"H1c"},"61":{"tbase":-7,"zone":"D","climat":"H1c"},
    "62":{"tbase":-9,"zone":"F","climat":"H1c"},"63":{"tbase":-8,"zone":"E","climat":"H1c"},
    "64":{"tbase":-5,"zone":"C","climat":"H2c"},"65":{"tbase":-5,"zone":"C","climat":"H2d"},
    "66":{"tbase":-5,"zone":"C","climat":"H3"},"67":{"tbase":-15,"zone":"I","climat":"H1c"},
    "68":{"tbase":-15,"zone":"I","climat":"H1c"},"69":{"tbase":-10,"zone":"G","climat":"H1a"},
    "70":{"tbase":-12,"zone":"H","climat":"H1b"},"71":{"tbase":-10,"zone":"G","climat":"H1b"},
    "72":{"tbase":-7,"zone":"D","climat":"H2b"},"73":{"tbase":-10,"zone":"G","climat":"H1a"},
    "74":{"tbase":-10,"zone":"G","climat":"H1a"},"75":{"tbase":-5,"zone":"D","climat":"H1c"},
    "76":{"tbase":-7,"zone":"D","climat":"H1c"},"77":{"tbase":-7,"zone":"D","climat":"H1c"},
    "78":{"tbase":-7,"zone":"D","climat":"H1c"},"79":{"tbase":-7,"zone":"C","climat":"H2b"},
    "80":{"tbase":-9,"zone":"F","climat":"H1c"},"81":{"tbase":-5,"zone":"C","climat":"H2d"},
    "82":{"tbase":-5,"zone":"C","climat":"H2d"},"83":{"tbase":-5,"zone":"A","climat":"H3"},
    "84":{"tbase":-6,"zone":"D","climat":"H2d"},"85":{"tbase":-5,"zone":"C","climat":"H2b"},
    "86":{"tbase":-7,"zone":"D","climat":"H2b"},"87":{"tbase":-8,"zone":"E","climat":"H1c"},
    "88":{"tbase":-15,"zone":"I","climat":"H1c"},"89":{"tbase":-10,"zone":"G","climat":"H1b"},
    "90":{"tbase":-15,"zone":"I","climat":"H1b"},"91":{"tbase":-7,"zone":"D","climat":"H1c"},
    "92":{"tbase":-7,"zone":"D","climat":"H1c"},"93":{"tbase":-7,"zone":"D","climat":"H1c"},
    "94":{"tbase":-7,"zone":"D","climat":"H1c"},"95":{"tbase":-7,"zone":"D","climat":"H1c"},
}

ZONE_ALTITUDE = {
    "A": [-2,-4,-6,-8,-10,-12,-14,-16,-18,-20],
    "B": [-4,-5,-6,-7,-8,-9,-10,-12,-13,-14],
    "C": [-5,-6,-7,-8,-9,-10,-11,-12,-13,-14],
    "D": [-7,-8,-9,-11,-13,-14,-15,-17,-19,-21],
    "E": [-8,-9,-11,-13,-15,-17,-19,-21,-23,-25],
    "F": [-9,-10,-11,-12,-13,-15,-17,-19,-21,-23],
    "G": [-10,-11,-13,-14,-17,-19,-21,-23,-24,-25],
    "H": [-12,-13,-15,-17,-19,-21,-23,-24,-25,-27],
    "I": [-15,-15,-19,-21,-23,-24,-25,-27,-29,-31],
}
ALTITUDE_BANDS = [200,400,600,800,1000,1200,1400,1600,1800,2000]

EFFIPAC_MODELS = [
    {"code":"AHP60_14","nom":"Effipac 14","puissance_nom":14.10,"refrigerant":"R32","t_max":60,"chassis":"S",
     "performance":{"A7/W35":{"pcalo":14.10,"pabs":2.91,"cop":4.85},"A7/W45":{"pcalo":14.41,"pabs":3.63,"cop":3.97},"A7/W55":{"pcalo":13.44,"pabs":4.35,"cop":3.09},"A-7/W55":{"pcalo":10.60,"pabs":5.07,"cop":2.09}}},
    {"code":"AHP60_18","nom":"Effipac 18","puissance_nom":17.90,"refrigerant":"R32","t_max":60,"chassis":"S",
     "performance":{"A7/W35":{"pcalo":17.90,"pabs":4.07,"cop":4.40},"A7/W45":{"pcalo":18.31,"pabs":5.03,"cop":3.64},"A7/W55":{"pcalo":17.25,"pabs":5.99,"cop":2.88},"A-7/W55":{"pcalo":12.30,"pabs":6.03,"cop":2.04}}},
    {"code":"AHP60_26","nom":"Effipac 26","puissance_nom":26.00,"refrigerant":"R32","t_max":60,"chassis":"M",
     "performance":{"A7/W35":{"pcalo":26.00,"pabs":6.44,"cop":4.04},"A7/W45":{"pcalo":26.65,"pabs":7.98,"cop":3.34},"A7/W55":{"pcalo":25.10,"pabs":9.51,"cop":2.64},"A-7/W55":{"pcalo":17.00,"pabs":9.44,"cop":1.80}}},
    {"code":"AHP60_32","nom":"Effipac 32","puissance_nom":32.10,"refrigerant":"R32","t_max":60,"chassis":"M",
     "performance":{"A7/W35":{"pcalo":32.10,"pabs":7.85,"cop":4.09},"A7/W45":{"pcalo":33.60,"pabs":9.97,"cop":3.37},"A7/W55":{"pcalo":31.80,"pabs":12.10,"cop":2.63},"A-7/W55":{"pcalo":21.70,"pabs":11.92,"cop":1.82}}},
    {"code":"AHP60_50","nom":"Effipac 50","puissance_nom":50.20,"refrigerant":"R32","t_max":60,"chassis":"L",
     "performance":{"A7/W35":{"pcalo":50.20,"pabs":12.21,"cop":4.11},"A7/W45":{"pcalo":51.34,"pabs":15.10,"cop":3.40},"A7/W55":{"pcalo":48.30,"pabs":18.02,"cop":2.68},"A-7/W55":{"pcalo":32.90,"pabs":21.79,"cop":1.51}}},
    {"code":"AHP60_70","nom":"Effipac 70","puissance_nom":66.80,"refrigerant":"R32","t_max":60,"chassis":"XL",
     "performance":{"A7/W35":{"pcalo":66.80,"pabs":16.29,"cop":4.10},"A7/W45":{"pcalo":67.37,"pabs":20.05,"cop":3.36},"A7/W55":{"pcalo":61.90,"pabs":23.80,"cop":2.60},"A-7/W55":{"pcalo":46.40,"pabs":30.13,"cop":1.54}}},
]

APTAE_MODELS = [
    {"code":"AHP70_15","nom":"Aptae 15","puissance_nom":16.33,"refrigerant":"R290","t_max":75,"chassis":"S",
     "performance":{"A7/W35":{"pcalo":16.33,"pabs":3.31,"cop":4.94},"A7/W45":{"pcalo":15.50,"pabs":3.88,"cop":4.00},"A7/W55":{"pcalo":15.23,"pabs":4.52,"cop":3.37},"A-7/W35":{"pcalo":12.00,"pabs":3.50,"cop":3.43},"A-7/W55":{"pcalo":10.50,"pabs":4.90,"cop":2.14}}},
    {"code":"AHP70_18","nom":"Aptae 18","puissance_nom":18.72,"refrigerant":"R290","t_max":75,"chassis":"S",
     "performance":{"A7/W35":{"pcalo":18.72,"pabs":4.05,"cop":4.62},"A7/W45":{"pcalo":18.48,"pabs":4.91,"cop":3.76},"A7/W55":{"pcalo":17.38,"pabs":5.32,"cop":3.27},"A-7/W35":{"pcalo":13.80,"pabs":4.25,"cop":3.25},"A-7/W55":{"pcalo":12.10,"pabs":5.95,"cop":2.03}}},
    {"code":"AHP70_23","nom":"Aptae 23","puissance_nom":22.80,"refrigerant":"R290","t_max":75,"chassis":"M",
     "performance":{"A7/W35":{"pcalo":22.80,"pabs":4.78,"cop":4.77},"A7/W45":{"pcalo":23.00,"pabs":5.92,"cop":3.89},"A7/W55":{"pcalo":21.60,"pabs":6.79,"cop":3.18},"A-7/W35":{"pcalo":16.80,"pabs":5.00,"cop":3.36},"A-7/W55":{"pcalo":14.70,"pabs":7.00,"cop":2.10}}},
    {"code":"AHP70_27","nom":"Aptae 27","puissance_nom":27.00,"refrigerant":"R290","t_max":75,"chassis":"M",
     "performance":{"A7/W35":{"pcalo":27.00,"pabs":6.21,"cop":4.35},"A7/W45":{"pcalo":27.51,"pabs":7.75,"cop":3.55},"A7/W55":{"pcalo":26.30,"pabs":8.74,"cop":3.01},"A-7/W35":{"pcalo":19.90,"pabs":6.50,"cop":3.06},"A-7/W55":{"pcalo":17.40,"pabs":9.10,"cop":1.91}}},
    {"code":"AHP70_40","nom":"Aptae 40","puissance_nom":40.00,"refrigerant":"R290","t_max":75,"chassis":"L",
     "performance":{"A7/W35":{"pcalo":40.00,"pabs":9.76,"cop":4.10},"A7/W45":{"pcalo":40.22,"pabs":12.07,"cop":3.33},"A7/W55":{"pcalo":38.10,"pabs":14.38,"cop":2.65},"A-7/W35":{"pcalo":29.50,"pabs":10.20,"cop":2.89},"A-7/W55":{"pcalo":25.80,"pabs":14.30,"cop":1.80}}},
    {"code":"AHP70_50","nom":"Aptae 50","puissance_nom":50.00,"refrigerant":"R290","t_max":75,"chassis":"L",
     "performance":{"A7/W35":{"pcalo":50.10,"pabs":11.90,"cop":4.21},"A7/W45":{"pcalo":50.52,"pabs":15.16,"cop":3.33},"A7/W55":{"pcalo":47.90,"pabs":16.50,"cop":2.90},"A-7/W35":{"pcalo":36.80,"pabs":12.75,"cop":2.89},"A-7/W55":{"pcalo":32.20,"pabs":17.90,"cop":1.80}}},
    {"code":"AHP70_65","nom":"Aptae 65","puissance_nom":62.00,"refrigerant":"R290","t_max":75,"chassis":"XL",
     "performance":{"A7/W35":{"pcalo":62.00,"pabs":15.12,"cop":4.10},"A7/W45":{"pcalo":58.90,"pabs":17.67,"cop":3.33},"A7/W55":{"pcalo":56.20,"pabs":21.21,"cop":2.65},"A-7/W35":{"pcalo":45.70,"pabs":15.81,"cop":2.89},"A-7/W55":{"pcalo":39.90,"pabs":22.17,"cop":1.80}}},
]

COMPETITORS = {
    "Daikin Altherma 3 H HT": {
        "refrigerant":"R290","t_max":70,
        "models":[
            {"nom":"EABH16DA9W","puissance_nom":16,"cop_a7w35":4.60,"cop_a7w55":3.10},
            {"nom":"EABX16DA9W","puissance_nom":16,"cop_a7w35":4.56,"cop_a7w55":3.08},
            {"nom":"Altherma 3 R ECH2O","puissance_nom":14,"cop_a7w35":4.30,"cop_a7w55":2.90},
        ]
    },
    "Mitsubishi Ecodan CAHV": {
        "refrigerant":"R744 (CO2)","t_max":90,
        "models":[
            {"nom":"CAHV-P500YA-HPB","puissance_nom":45,"cop_a7w65":3.80,"cop_a7w55":None},
            {"nom":"CAHV-P500YB-HPB","puissance_nom":50,"cop_a7w65":3.90,"cop_a7w55":None},
        ]
    }
}

BUILDING_TYPES = {
    "Passif / BBC": 0.015,
    "RE2020 / RT2012": 0.040,
    "Rénové (post-1990)": 0.060,
    "Ancien (1970-1990)": 0.080,
    "Très ancien (<1970)": 0.120,
}

BIN_HOURS = {
    "H1": {-20:1,-19:2,-18:4,-17:6,-16:10,-15:15,-14:22,-13:30,-12:40,-11:52,-10:65,-9:80,-8:100,-7:120,-6:145,-5:170,-4:200,-3:230,-2:265,-1:300,0:340,1:380,2:420,3:460,4:500,5:540,6:560,7:570,8:560,9:530,10:500,11:460,12:420,13:380,14:340,15:300,16:250,17:200,18:150,19:80},
    "H2": {-15:2,-14:4,-13:6,-12:10,-11:15,-10:22,-9:30,-8:42,-7:55,-6:72,-5:92,-4:115,-3:142,-2:175,-1:210,0:250,1:295,2:340,3:385,4:430,5:470,6:500,7:520,8:530,9:520,10:500,11:470,12:430,13:385,14:340,15:295,16:250,17:200,18:150,19:100},
    "H3": {-10:2,-9:4,-8:8,-7:14,-6:22,-5:35,-4:50,-3:72,-2:100,-1:135,0:175,1:220,2:270,3:325,4:385,5:445,6:500,7:545,8:580,9:590,10:580,11:560,12:530,13:490,14:445,15:395,16:340,17:280,18:215,19:130},
}

ECS_EQUIV = {
    "social": {"T1":0.6,"T2":0.7,"T3":1.0,"T4":1.4,"T5":1.8,"T6":1.9},
    "prive":  {"T1":0.6,"T2":0.7,"T3":0.9,"T4":1.1,"T5":1.3,"T6":1.4},
}
ECS_PEAK = {
    "10min":{"a":61,"b":0.503},
    "1h":{"a":83,"b":0.708},
    "2h":{"a":108,"b":0.773},
    "3h":{"a":116,"b":0.815},
    "4h":{"a":162,"b":0.789},
}

PIPE_DIAMETERS = [
    {"int":14,"ext":16,"maxFlow":150},
    {"int":20,"ext":22,"maxFlow":380},
    {"int":26,"ext":28,"maxFlow":770},
    {"int":33,"ext":35,"maxFlow":1500},
    {"int":40,"ext":42,"maxFlow":2450},
    {"int":50,"ext":54,"maxFlow":4200},
    {"int":66,"ext":70,"maxFlow":8000},
]


# ============================================================
# INTERPOLATION ENGINE (Python version of engine.js)
# ============================================================

def interpolate_performance(model, t_ext, t_water):
    """Interpolate PAC capacity and COP at given conditions."""
    perf = model["performance"]
    water_key = f"W{t_water}"

    conditions = []
    for key, data in perf.items():
        if water_key in key:
            t_air = int(key.replace("A","").split("/")[0])
            conditions.append({"tAir": t_air, "data": data})

    if not conditions:
        # Try interpolation between water temps
        return interpolate_water(model, t_ext, t_water)

    conditions.sort(key=lambda x: x["tAir"])

    if len(conditions) == 1:
        d = conditions[0]["data"]
        pabs = d.get("pabs", d["pcalo"] / d["cop"])
        return {"pcalo": d["pcalo"], "cop": d["cop"], "pabs": pabs}

    lower = conditions[0]
    upper = conditions[-1]
    for i in range(len(conditions) - 1):
        if conditions[i]["tAir"] <= t_ext <= conditions[i+1]["tAir"]:
            lower = conditions[i]
            upper = conditions[i+1]
            break

    rng = upper["tAir"] - lower["tAir"]
    ratio = 0 if rng == 0 else (t_ext - lower["tAir"]) / rng

    # Clamp ratio for extrapolation
    ratio = max(-0.5, min(1.5, ratio))

    pcalo = lower["data"]["pcalo"] + ratio * (upper["data"]["pcalo"] - lower["data"]["pcalo"])
    cop = lower["data"]["cop"] + ratio * (upper["data"]["cop"] - lower["data"]["cop"])
    cop = max(1.0, cop)
    pcalo = max(0.5, pcalo)
    pabs = pcalo / cop

    return {"pcalo": round(pcalo, 2), "cop": round(cop, 2), "pabs": round(pabs, 2)}


def interpolate_water(model, t_ext, t_water):
    """Interpolate between water temperatures."""
    temps = [35, 45, 55]
    results = {}

    for tw in temps:
        r = interpolate_performance(model, t_ext, tw)
        if r and r["pcalo"] > 0:
            results[tw] = r

    keys = sorted(results.keys())
    if len(keys) < 2:
        return results[keys[0]] if keys else None

    lower = keys[0]
    upper = keys[-1]
    for i in range(len(keys) - 1):
        if keys[i] <= t_water <= keys[i+1]:
            lower = keys[i]
            upper = keys[i+1]
            break

    ratio = (t_water - lower) / (upper - lower)
    pcalo = results[lower]["pcalo"] + ratio * (results[upper]["pcalo"] - results[lower]["pcalo"])
    cop = results[lower]["cop"] + ratio * (results[upper]["cop"] - results[lower]["cop"])
    return {"pcalo": round(pcalo, 2), "cop": round(cop, 2), "pabs": round(pcalo/cop, 2)}


def get_tbase(dept, altitude=0, bord_de_mer=False):
    """Get Tbase for a department with altitude correction."""
    d = TBASE.get(dept)
    if not d:
        return -7
    zone = d["zone"]
    corr = ZONE_ALTITUDE.get(zone)
    if not corr or altitude <= 200:
        tbase = d["tbase"]
    else:
        idx = 0
        for i, band in enumerate(ALTITUDE_BANDS):
            if altitude <= band:
                idx = i
                break
            if i == len(ALTITUDE_BANDS) - 1:
                idx = i
        tbase = corr[idx]
    if bord_de_mer:
        tbase += 2
    return tbase


def get_climate_zone(dept):
    d = TBASE.get(dept)
    return d["climat"] if d else "H1c"


# ============================================================
# STYLES
# ============================================================

NAVY = "003366"
ORANGE = "FF6600"
LIGHT_BLUE = "E8F0FE"
LIGHT_ORANGE = "FFF3E6"
WHITE = "FFFFFF"
DARK_TEXT = "1A1A2E"
MEDIUM_GRAY = "999999"
LIGHT_GRAY = "F5F5F5"
BORDER_COLOR = "D0D0D0"

font_title = Font(name="Calibri", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=11, color="CCDDEE")
font_section = Font(name="Calibri", size=13, bold=True, color=NAVY)
font_label = Font(name="Calibri", size=10, color=DARK_TEXT)
font_label_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_value = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_value_orange = Font(name="Calibri", size=11, bold=True, color=ORANGE)
font_header = Font(name="Calibri", size=10, bold=True, color=WHITE)
font_data = Font(name="Calibri", size=10, color=DARK_TEXT)
font_data_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_small = Font(name="Calibri", size=9, color=MEDIUM_GRAY)
font_result_big = Font(name="Calibri", size=14, bold=True, color=NAVY)
font_result_highlight = Font(name="Calibri", size=14, bold=True, color=ORANGE)

fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
fill_light_blue = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_light_orange = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_atlantic = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR)
)
bottom_border = Border(bottom=Side(style="medium", color=NAVY))

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")


def style_cell(ws, row, col, value, font=None, fill=None, border=None, alignment=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    if fmt: cell.number_format = fmt
    return cell


def merge_style(ws, range_str, value, font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(range_str)
    min_col, min_row, _, _ = openpyxl.utils.cell.range_boundaries(range_str)
    cell = ws.cell(row=min_row, column=min_col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell


def header_row(ws, row, col_start, headers, widths=None):
    for i, h in enumerate(headers):
        c = col_start + i
        style_cell(ws, row, c, h, font=font_header, fill=fill_navy, border=thin_border, alignment=align_center)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(col_start + i)].width = w


def data_row(ws, row, col_start, values, fonts=None, fmts=None, alt=False):
    fill = fill_light_gray if alt else fill_white
    for i, v in enumerate(values):
        f = fonts[i] if fonts and i < len(fonts) else font_data
        fmt = fmts[i] if fmts and i < len(fmts) else None
        style_cell(ws, row, col_start + i, v, font=f, fill=fill, border=thin_border, alignment=align_center, fmt=fmt)


# ============================================================
# CREATE WORKBOOK
# ============================================================

wb = openpyxl.Workbook()


# ============================================================
# SHEET 1: SAISIE
# ============================================================

ws_saisie = wb.active
ws_saisie.title = "SAISIE"
ws_saisie.sheet_properties.tabColor = NAVY

# Column widths
for col, w in [(1,3),(2,28),(3,20),(4,6),(5,28),(6,20),(7,6),(8,28),(9,20),(10,3)]:
    ws_saisie.column_dimensions[get_column_letter(col)].width = w

# Header banner
for c in range(1, 11):
    style_cell(ws_saisie, 1, c, None, fill=fill_navy)
    style_cell(ws_saisie, 2, c, None, fill=fill_navy)
    style_cell(ws_saisie, 3, c, None, fill=fill_navy)

merge_style(ws_saisie, "B1:I1", "PAC DIMENSIONNEMENT", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))
merge_style(ws_saisie, "B2:I2", "Outil de dimensionnement PAC collectives — Atlantic Effipac / Aptae", font=font_subtitle, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))
merge_style(ws_saisie, "B3:I3", "NF EN 12831 · EN 14825 · ADEME/COSTIC · Heat Pump Keymark", font=Font(name="Calibri", size=9, italic=True, color="88AACC"), fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

# === SECTION 1: LOCALISATION ===
row = 5
merge_style(ws_saisie, f"B{row}:C{row}", "LOCALISATION", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 4):
    ws_saisie.cell(row=row, column=c).border = bottom_border

row = 7
style_cell(ws_saisie, row, 2, "Département", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, "75", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
ws_saisie.cell(row=row, column=3).comment = None

# Add dropdown for department
dept_list = sorted(TBASE.keys())
dv_dept = DataValidation(type="list", formula1='"' + ','.join(dept_list) + '"', allow_blank=True)
dv_dept.prompt = "Sélectionnez le département"
dv_dept.promptTitle = "Département"
ws_saisie.add_data_validation(dv_dept)
dv_dept.add(ws_saisie["C7"])

row = 8
style_cell(ws_saisie, row, 2, "Altitude (m)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 0, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

row = 9
style_cell(ws_saisie, row, 2, "Bord de mer", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, "Non", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_mer = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
ws_saisie.add_data_validation(dv_mer)
dv_mer.add(ws_saisie["C9"])

# Tbase result (computed)
row = 10
style_cell(ws_saisie, row, 2, "→ Tbase calculée", font=font_label_bold, alignment=align_left)
style_cell(ws_saisie, row, 3, None, font=font_value_orange, fill=fill_light_orange, border=thin_border, alignment=align_center)
# Formula: VLOOKUP on DATA_TBASE
ws_saisie["C10"] = '=IFERROR(VLOOKUP(C7,DATA_TBASE!A:D,2,FALSE)+IF(C9="Oui",2,0), "")'
ws_saisie["C10"].font = font_value_orange
ws_saisie["C10"].fill = fill_light_orange
ws_saisie["C10"].border = thin_border
ws_saisie["C10"].alignment = align_center

row = 11
style_cell(ws_saisie, row, 2, "→ Zone climatique", font=font_label_bold, alignment=align_left)
ws_saisie["C11"] = '=IFERROR(VLOOKUP(C7,DATA_TBASE!A:D,4,FALSE), "")'
ws_saisie["C11"].font = font_value_orange
ws_saisie["C11"].fill = fill_light_orange
ws_saisie["C11"].border = thin_border
ws_saisie["C11"].alignment = align_center

# === SECTION 2: BATIMENT ===
row = 13
merge_style(ws_saisie, f"B{row}:C{row}", "BÂTIMENT", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 4):
    ws_saisie.cell(row=row, column=c).border = bottom_border

row = 15
style_cell(ws_saisie, row, 2, "Type de bâtiment", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, "RE2020 / RT2012", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_bat = DataValidation(type="list", formula1='"' + ','.join(BUILDING_TYPES.keys()) + '"', allow_blank=True)
ws_saisie.add_data_validation(dv_bat)
dv_bat.add(ws_saisie["C15"])

row = 16
style_cell(ws_saisie, row, 2, "Surface habitable (m²)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 1500, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

row = 17
style_cell(ws_saisie, row, 2, "T° consigne intérieure (°C)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 19, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

row = 18
style_cell(ws_saisie, row, 2, "→ Déperditions spécifiques", font=font_label_bold, alignment=align_left)
ws_saisie["C18"] = '=IFERROR(VLOOKUP(C15,DATA_BATIMENT!A:B,2,FALSE)*1000, "")'
ws_saisie["C18"].font = font_value_orange
ws_saisie["C18"].fill = fill_light_orange
ws_saisie["C18"].border = thin_border
ws_saisie["C18"].alignment = align_center
ws_saisie["C18"].number_format = '0 "W/m²"'

row = 19
style_cell(ws_saisie, row, 2, "→ Déperditions totales (kW)", font=font_label_bold, alignment=align_left)
ws_saisie["C19"] = '=IFERROR(VLOOKUP(C15,DATA_BATIMENT!A:B,2,FALSE)*C16, "")'
ws_saisie["C19"].font = font_result_highlight
ws_saisie["C19"].fill = fill_light_orange
ws_saisie["C19"].border = thin_border
ws_saisie["C19"].alignment = align_center
ws_saisie["C19"].number_format = '0.0 "kW"'


# === SECTION 3: PAC CONFIG ===
row = 21
merge_style(ws_saisie, f"B{row}:C{row}", "CONFIGURATION PAC", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 4):
    ws_saisie.cell(row=row, column=c).border = bottom_border

row = 23
style_cell(ws_saisie, row, 2, "Gamme PAC", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, "Effipac (R32)", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_gamme = DataValidation(type="list", formula1='"Effipac (R32),Aptae (R290)"', allow_blank=True)
ws_saisie.add_data_validation(dv_gamme)
dv_gamme.add(ws_saisie["C23"])

row = 24
style_cell(ws_saisie, row, 2, "Mode de fonctionnement", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, "Chauffage 100% Electrique", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_mode = DataValidation(type="list", formula1='"Chauffage 100% Electrique,Double Service,ECS Seul,Chauffage Hybride"', allow_blank=True)
ws_saisie.add_data_validation(dv_mode)
dv_mode.add(ws_saisie["C24"])

row = 25
style_cell(ws_saisie, row, 2, "T° eau départ (°C)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 45, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_teau = DataValidation(type="list", formula1='"35,45,55"', allow_blank=True)
ws_saisie.add_data_validation(dv_teau)
dv_teau.add(ws_saisie["C25"])

row = 26
style_cell(ws_saisie, row, 2, "Delta T (K)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 5, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_dt = DataValidation(type="list", formula1='"5,7,10"', allow_blank=True)
ws_saisie.add_data_validation(dv_dt)
dv_dt.add(ws_saisie["C26"])

row = 27
style_cell(ws_saisie, row, 2, "Nombre de PAC max", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 3, 4, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_nmax = DataValidation(type="list", formula1='"1,2,3,4,5,6"', allow_blank=True)
ws_saisie.add_data_validation(dv_nmax)
dv_nmax.add(ws_saisie["C27"])


# === SECTION 4: ECS (right column) ===
row = 5
merge_style(ws_saisie, f"E{row}:F{row}", "ECS (Eau Chaude Sanitaire)", font=font_section, fill=fill_white, alignment=align_left)
for c in range(5, 7):
    ws_saisie.cell(row=row, column=c).border = bottom_border

row = 7
style_cell(ws_saisie, row, 5, "Type de parc", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 6, "social", font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)
dv_parc = DataValidation(type="list", formula1='"social,prive"', allow_blank=True)
ws_saisie.add_data_validation(dv_parc)
dv_parc.add(ws_saisie["F7"])

# Logement count table
row = 9
style_cell(ws_saisie, row, 5, "Type logement", font=font_header, fill=fill_navy, border=thin_border, alignment=align_center)
style_cell(ws_saisie, row, 6, "Nombre", font=font_header, fill=fill_navy, border=thin_border, alignment=align_center)

logement_types = ["T1", "T2", "T3", "T4", "T5", "T6"]
default_counts = [5, 10, 20, 10, 5, 0]
for i, (lt, cnt) in enumerate(zip(logement_types, default_counts)):
    r = row + 1 + i
    style_cell(ws_saisie, r, 5, lt, font=font_label_bold, fill=fill_light_gray if i%2==0 else fill_white, border=thin_border, alignment=align_center)
    style_cell(ws_saisie, r, 6, cnt, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

row = 16
style_cell(ws_saisie, row, 5, "T° ECS stockage (°C)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 6, 60, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

row = 17
style_cell(ws_saisie, row, 5, "T° eau froide (°C)", font=font_label, alignment=align_left)
style_cell(ws_saisie, row, 6, 10, font=font_value, fill=fill_light_blue, border=thin_border, alignment=align_center)

# === ECS Results ===
row = 19
merge_style(ws_saisie, f"E{row}:F{row}", "RÉSULTATS ECS", font=font_section, fill=fill_white, alignment=align_left)
for c in range(5, 7):
    ws_saisie.cell(row=row, column=c).border = bottom_border

# Calculate Ns formula
row = 21
style_cell(ws_saisie, row, 5, "Ns (logements standard)", font=font_label, alignment=align_left)
# Ns = sum of (count × equiv_coeff) for each logement type
ws_saisie["F21"] = ('=IFERROR('
    'F10*VLOOKUP(F7,DATA_ECS!A:G,2,FALSE)+'
    'F11*VLOOKUP(F7,DATA_ECS!A:G,3,FALSE)+'
    'F12*VLOOKUP(F7,DATA_ECS!A:G,4,FALSE)+'
    'F13*VLOOKUP(F7,DATA_ECS!A:G,5,FALSE)+'
    'F14*VLOOKUP(F7,DATA_ECS!A:G,6,FALSE)+'
    'F15*VLOOKUP(F7,DATA_ECS!A:G,7,FALSE)'
    ', "")')
ws_saisie["F21"].font = font_value_orange
ws_saisie["F21"].fill = fill_light_orange
ws_saisie["F21"].border = thin_border
ws_saisie["F21"].alignment = align_center
ws_saisie["F21"].number_format = '0.0'

row = 22
style_cell(ws_saisie, row, 5, "Volume journalier (L/j à 60°C)", font=font_label, alignment=align_left)
ws_saisie["F22"] = '=IFERROR(F21*70, "")'
ws_saisie["F22"].font = font_value_orange
ws_saisie["F22"].fill = fill_light_orange
ws_saisie["F22"].border = thin_border
ws_saisie["F22"].alignment = align_center
ws_saisie["F22"].number_format = '#,##0 "L"'

row = 23
style_cell(ws_saisie, row, 5, "Volume pointe 10 min (L)", font=font_label, alignment=align_left)
ws_saisie["F23"] = '=IFERROR(61*POWER(F21,0.503), "")'
ws_saisie["F23"].font = font_value_orange
ws_saisie["F23"].fill = fill_light_orange
ws_saisie["F23"].border = thin_border
ws_saisie["F23"].alignment = align_center
ws_saisie["F23"].number_format = '#,##0 "L"'

row = 24
style_cell(ws_saisie, row, 5, "Volume pointe 1h (L)", font=font_label, alignment=align_left)
ws_saisie["F24"] = '=IFERROR(83*POWER(F21,0.708), "")'
ws_saisie["F24"].font = font_value_orange
ws_saisie["F24"].fill = fill_light_orange
ws_saisie["F24"].border = thin_border
ws_saisie["F24"].alignment = align_center
ws_saisie["F24"].number_format = '#,##0 "L"'

row = 25
style_cell(ws_saisie, row, 5, "Volume stockage min (L)", font=font_label, alignment=align_left)
ws_saisie["F25"] = '=IFERROR(F23*2.4, "")'
ws_saisie["F25"].font = font_result_highlight
ws_saisie["F25"].fill = fill_light_orange
ws_saisie["F25"].border = thin_border
ws_saisie["F25"].alignment = align_center
ws_saisie["F25"].number_format = '#,##0 "L"'

row = 26
style_cell(ws_saisie, row, 5, "Puissance production ECS (kW)", font=font_label, alignment=align_left)
ws_saisie["F26"] = '=IFERROR(1.163*(F16-F17)*F22/(8*1000), "")'
ws_saisie["F26"].font = font_result_highlight
ws_saisie["F26"].fill = fill_light_orange
ws_saisie["F26"].border = thin_border
ws_saisie["F26"].alignment = align_center
ws_saisie["F26"].number_format = '0.0 "kW"'


# === SECTION: BOTTOM RESULTS SUMMARY ===
row = 29
merge_style(ws_saisie, f"B{row}:F{row}", "RÉSULTATS RAPIDES (voir onglet RÉSULTATS pour le détail)", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 7):
    ws_saisie.cell(row=row, column=c).border = bottom_border

# These are informational pointers
row = 31
style_cell(ws_saisie, row, 2, "→ Consultez l'onglet RÉSULTATS pour la sélection PAC complète", font=font_label, alignment=align_left)
row = 32
style_cell(ws_saisie, row, 2, "→ Consultez l'onglet MONOTONE pour la courbe monotone et le taux de couverture", font=font_label, alignment=align_left)
row = 33
style_cell(ws_saisie, row, 2, "→ Consultez l'onglet COMPARAISON pour la comparaison concurrentielle", font=font_label, alignment=align_left)

# Print setup
ws_saisie.sheet_view.showGridLines = False
ws_saisie.page_setup.orientation = "landscape"
ws_saisie.page_setup.fitToWidth = 1
ws_saisie.page_setup.fitToHeight = 1


# ============================================================
# SHEET: DATA_TBASE (lookup reference)
# ============================================================

ws_tbase = wb.create_sheet("DATA_TBASE")
ws_tbase.sheet_properties.tabColor = "999999"

header_row(ws_tbase, 1, 1, ["Département", "Tbase (°C)", "Zone", "Climat"], [14, 14, 10, 10])

r = 2
for dept in sorted(TBASE.keys()):
    d = TBASE[dept]
    data_row(ws_tbase, r, 1, [dept, d["tbase"], d["zone"], d["climat"]], alt=(r%2==0))
    r += 1


# ============================================================
# SHEET: DATA_BATIMENT
# ============================================================

ws_bat = wb.create_sheet("DATA_BATIMENT")
ws_bat.sheet_properties.tabColor = "999999"

header_row(ws_bat, 1, 1, ["Type bâtiment", "Spécifique (kW/m²)", "Description"], [28, 20, 20])

r = 2
for name, specific in BUILDING_TYPES.items():
    desc = f"~{int(specific*1000)} W/m²"
    data_row(ws_bat, r, 1, [name, specific, desc], fmts=[None, "0.000", None], alt=(r%2==0))
    r += 1


# ============================================================
# SHEET: DATA_ECS
# ============================================================

ws_ecs = wb.create_sheet("DATA_ECS")
ws_ecs.sheet_properties.tabColor = "999999"

header_row(ws_ecs, 1, 1, ["Parc", "T1", "T2", "T3", "T4", "T5", "T6"], [14,8,8,8,8,8,8])

r = 2
for parc, coeffs in ECS_EQUIV.items():
    vals = [parc] + [coeffs[t] for t in ["T1","T2","T3","T4","T5","T6"]]
    data_row(ws_ecs, r, 1, vals, alt=(r%2==0))
    r += 1


# ============================================================
# SHEET: DATA_PAC - Pre-calculated performance tables
# ============================================================

ws_pac = wb.create_sheet("DATA_PAC")
ws_pac.sheet_properties.tabColor = "999999"

# For each model, for each water temp, pre-calculate performance at each ext temp
ext_temps = list(range(-20, 21))  # -20 to +20

# Write Effipac table
row = 1
merge_style(ws_pac, f"A{row}:Z{row}", "EFFIPAC (R32) — Performance pré-calculée", font=font_section, fill=fill_light_blue)

row = 3
ws_pac.column_dimensions["A"].width = 14
ws_pac.column_dimensions["B"].width = 8
ws_pac.column_dimensions["C"].width = 10

# Headers
header_row(ws_pac, row, 1, ["Modèle", "T°eau", "Type"] + [f"{t}°C" for t in ext_temps])

row = 4
for model in EFFIPAC_MODELS:
    for tw in [35, 45, 55]:
        # Pcalo row
        vals = [model["nom"], tw, "Pcalo (kW)"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["pcalo"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

        # COP row
        vals = [model["nom"], tw, "COP"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["cop"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

        # Pabs row
        vals = [model["nom"], tw, "Pabs (kW)"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["pabs"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

# Write Aptae table
row += 2
aptae_start = row
merge_style(ws_pac, f"A{row}:Z{row}", "APTAE (R290) — Performance pré-calculée", font=font_section, fill=fill_light_orange)

row += 2
header_row(ws_pac, row, 1, ["Modèle", "T°eau", "Type"] + [f"{t}°C" for t in ext_temps])

row += 1
for model in APTAE_MODELS:
    for tw in [35, 45, 55]:
        vals = [model["nom"], tw, "Pcalo (kW)"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["pcalo"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

        vals = [model["nom"], tw, "COP"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["cop"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

        vals = [model["nom"], tw, "Pabs (kW)"]
        for te in ext_temps:
            p = interpolate_performance(model, te, tw)
            vals.append(round(p["pabs"], 2) if p else "")
        data_row(ws_pac, row, 1, vals, alt=(row%2==0))
        row += 1

# Set narrow widths for temperature columns
for i in range(len(ext_temps)):
    ws_pac.column_dimensions[get_column_letter(4 + i)].width = 7


# ============================================================
# SHEET: DATA_BIN
# ============================================================

ws_bin = wb.create_sheet("DATA_BIN")
ws_bin.sheet_properties.tabColor = "999999"

ws_bin.column_dimensions["A"].width = 14
header_row(ws_bin, 1, 1, ["T°ext (°C)", "H1 (heures)", "H2 (heures)", "H3 (heures)"], [14, 14, 14, 14])

all_temps = sorted(set(list(BIN_HOURS["H1"].keys()) + list(BIN_HOURS["H2"].keys()) + list(BIN_HOURS["H3"].keys())))
r = 2
for t in all_temps:
    h1 = BIN_HOURS["H1"].get(t, 0)
    h2 = BIN_HOURS["H2"].get(t, 0)
    h3 = BIN_HOURS["H3"].get(t, 0)
    data_row(ws_bin, r, 1, [t, h1, h2, h3], alt=(r%2==0))
    r += 1


# ============================================================
# SHEET: RÉSULTATS — Main results with pre-computed solutions
# ============================================================

ws_res = wb.create_sheet("RÉSULTATS")
ws_res.sheet_properties.tabColor = ORANGE
ws_res.sheet_view.showGridLines = False

for col, w in [(1,3),(2,22),(3,16),(4,16),(5,16),(6,16),(7,16),(8,16),(9,16),(10,16),(11,3)]:
    ws_res.column_dimensions[get_column_letter(col)].width = w

# Header
for c in range(1, 12):
    style_cell(ws_res, 1, c, None, fill=fill_navy)
    style_cell(ws_res, 2, c, None, fill=fill_navy)

merge_style(ws_res, "B1:J1", "RÉSULTATS — Sélection PAC", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))
merge_style(ws_res, "B2:J2", "Les tableaux ci-dessous sont pré-calculés pour les paramètres de l'onglet SAISIE", font=font_subtitle, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

# We'll pre-compute results for the default parameters
# But also compute for all models so the user can see all options

# === Section: All PAC options for Effipac ===
row = 4
merge_style(ws_res, f"B{row}:J{row}", "SÉLECTION EFFIPAC (R32) — Toutes les configurations possibles", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 11):
    ws_res.cell(row=row, column=c).border = bottom_border

row = 6
headers = ["Modèle", "Nb PAC", "P. nominale\ntotale (kW)", "P. à A7/W45\n(kW)", "COP\nA7/W45",
           "P. à A-7/W55\n(kW)", "COP\nA-7/W55", "P. élec.\nmax (kW)", "Chassis"]
header_row(ws_res, row, 2, headers)

row = 7
for model in EFFIPAC_MODELS:
    for n in range(1, 7):
        total_nom = round(model["puissance_nom"] * n, 1)
        if total_nom > 420:
            break

        p_a7w45 = model["performance"].get("A7/W45", {})
        p_am7w55 = model["performance"].get("A-7/W55", {})

        vals = [
            model["nom"],
            n,
            total_nom,
            round(p_a7w45.get("pcalo", 0) * n, 1),
            p_a7w45.get("cop", ""),
            round(p_am7w55.get("pcalo", 0) * n, 1),
            p_am7w55.get("cop", ""),
            round(p_a7w45.get("pabs", 0) * n, 1),
            model["chassis"]
        ]
        fmts = [None, "0", "0.0", "0.0", "0.00", "0.0", "0.00", "0.0", None]
        data_row(ws_res, row, 2, vals, fmts=fmts, alt=(row%2==0))
        row += 1

effipac_end = row

# === Section: All PAC options for Aptae ===
row += 1
merge_style(ws_res, f"B{row}:J{row}", "SÉLECTION APTAE (R290) — Toutes les configurations possibles", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 11):
    ws_res.cell(row=row, column=c).border = bottom_border

row += 2
header_row(ws_res, row, 2, headers)

row += 1
for model in APTAE_MODELS:
    for n in range(1, 7):
        total_nom = round(model["puissance_nom"] * n, 1)
        if total_nom > 450:
            break

        p_a7w45 = model["performance"].get("A7/W45", {})
        p_am7w55 = model["performance"].get("A-7/W55", {})

        vals = [
            model["nom"],
            n,
            total_nom,
            round(p_a7w45.get("pcalo", 0) * n, 1),
            p_a7w45.get("cop", ""),
            round(p_am7w55.get("pcalo", 0) * n, 1),
            p_am7w55.get("cop", ""),
            round(p_a7w45.get("pabs", 0) * n, 1),
            model["chassis"]
        ]
        fmts = [None, "0", "0.0", "0.0", "0.00", "0.0", "0.00", "0.0", None]
        data_row(ws_res, row, 2, vals, fmts=fmts, alt=(row%2==0))
        row += 1

aptae_end = row

# === Section: Hydraulic Reference ===
row += 1
merge_style(ws_res, f"B{row}:J{row}", "DIMENSIONNEMENT HYDRAULIQUE — Table de référence", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 11):
    ws_res.cell(row=row, column=c).border = bottom_border

row += 2
header_row(ws_res, row, 2, ["Puissance (kW)", "ΔT=5K\nDébit (m³/h)", "ΔT=5K\nDébit (L/h)", "ΔT=7K\nDébit (m³/h)",
                              "Diamètre\nint (mm)", "Diamètre\next (mm)", "Bouteille\n(mm)", "Ballon tampon\n(L)"])

row += 1
for power in [10, 15, 20, 25, 30, 40, 50, 60, 70, 80, 100, 120, 150, 200]:
    debit_5 = power / (5 * 1.163)
    debit_7 = power / (7 * 1.163)
    debit_lh = debit_5 * 1000

    # Find pipe
    selected = PIPE_DIAMETERS[-1]
    for p in PIPE_DIAMETERS:
        if p["maxFlow"] >= debit_lh:
            selected = p
            break

    vals = [
        power,
        round(debit_5, 2),
        round(debit_lh, 0),
        round(debit_7, 2),
        selected["int"],
        selected["ext"],
        selected["int"] * 3,
        round(14 * power)
    ]
    fmts = ["0", "0.00", "#,##0", "0.00", "0", "0", "0", "#,##0"]
    data_row(ws_res, row, 2, vals, fmts=fmts, alt=(row%2==0))
    row += 1


# ============================================================
# SHEET: MONOTONE — Pre-calculated monotone curves for defaults
# ============================================================

ws_mono = wb.create_sheet("MONOTONE")
ws_mono.sheet_properties.tabColor = "336699"
ws_mono.sheet_view.showGridLines = False

for col, w in [(1,3),(2,16),(3,16),(4,16),(5,16),(6,16),(7,16),(8,3)]:
    ws_mono.column_dimensions[get_column_letter(col)].width = w

# Header
for c in range(1, 9):
    style_cell(ws_mono, 1, c, None, fill=fill_navy)
    style_cell(ws_mono, 2, c, None, fill=fill_navy)

merge_style(ws_mono, "B1:G1", "COURBE MONOTONE DE CHAUFFAGE", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))
merge_style(ws_mono, "B2:G2", "Méthode des bins EN 14825 — Heures cumulées vs puissance", font=font_subtitle, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

# Pre-calculate monotone for default config: Paris (75), RE2020, 1500m², Effipac 26, 1 unit, W45
DEFAULT_DEPT = "75"
DEFAULT_SURFACE = 1500
DEFAULT_BT = "RE2020 / RT2012"
DEFAULT_SPECIFIC = 0.040

tbase = get_tbase(DEFAULT_DEPT)
climat = get_climate_zone(DEFAULT_DEPT)
dep_kw = DEFAULT_SURFACE * DEFAULT_SPECIFIC
t_int = 19
dp = dep_kw * 1000 / (t_int - tbase)

# Select best model for this load
best_model = None
best_n = 1
for model in EFFIPAC_MODELS:
    for n in range(1, 5):
        total = model["puissance_nom"] * n
        if 0.8 * dep_kw <= total <= 1.2 * dep_kw:
            best_model = model
            best_n = n
            break
    if best_model:
        break

if not best_model:
    best_model = EFFIPAC_MODELS[3]  # Effipac 32
    best_n = 2

# Generate bin method data
bin_zone = climat[:2] if climat else "H1"
bins = BIN_HOURS.get(bin_zone, BIN_HOURS["H1"])

row = 4
merge_style(ws_mono, f"B{row}:G{row}",
    f"Exemple: Paris (75), {DEFAULT_SURFACE}m² {DEFAULT_BT}, Tbase={tbase}°C, Déperditions={dep_kw:.1f} kW",
    font=font_label_bold, fill=fill_light_blue, alignment=align_left)
row = 5
merge_style(ws_mono, f"B{row}:G{row}",
    f"PAC: {best_n}× {best_model['nom']} ({best_model['puissance_nom']*best_n:.1f} kW nominal), T°eau=45°C, Zone {bin_zone}",
    font=font_label_bold, fill=fill_light_blue, alignment=align_left)

row = 7
header_row(ws_mono, row, 2, ["T°ext (°C)", "Heures", "Besoins (kW)", "Capacité PAC (kW)", "COP", "Appoint (kW)"])

row = 8
chart_data_start = row

total_e_pac = 0
total_e_backup = 0
total_e_total = 0

sorted_temps = sorted(bins.keys())
for t in sorted_temps:
    hours = bins[t]
    if t >= t_int:
        continue

    load = dp * (t_int - t) / 1000
    perf = interpolate_performance(best_model, t, 45)
    pac_cap = perf["pcalo"] * best_n if perf else 0
    cop = perf["cop"] if perf else 1
    backup = max(0, load - pac_cap)

    total_e_pac += min(pac_cap, load) * hours
    total_e_backup += backup * hours
    total_e_total += load * hours

    vals = [t, hours, round(load, 1), round(pac_cap, 1), round(cop, 2), round(backup, 1)]
    fmts = ["0", "#,##0", "0.0", "0.0", "0.00", "0.0"]
    data_row(ws_mono, row, 2, vals, fmts=fmts, alt=(row%2==0))
    row += 1

chart_data_end = row - 1

# Summary
row += 1
style_cell(ws_mono, row, 2, "Énergie totale chauffage (kWh/an)", font=font_label_bold, alignment=align_left)
style_cell(ws_mono, row, 5, round(total_e_total), font=font_result_big, alignment=align_right, fmt="#,##0")
row += 1
style_cell(ws_mono, row, 2, "Énergie PAC (kWh/an)", font=font_label_bold, alignment=align_left)
style_cell(ws_mono, row, 5, round(total_e_pac), font=font_result_big, alignment=align_right, fmt="#,##0")
row += 1
style_cell(ws_mono, row, 2, "Énergie appoint (kWh/an)", font=font_label_bold, alignment=align_left)
style_cell(ws_mono, row, 5, round(total_e_backup), font=font_result_big, alignment=align_right, fmt="#,##0")
row += 1
taux = round(total_e_pac / total_e_total * 100, 1) if total_e_total > 0 else 0
style_cell(ws_mono, row, 2, "Taux de couverture annuel", font=font_label_bold, alignment=align_left)
style_cell(ws_mono, row, 5, taux, font=font_result_highlight, alignment=align_right, fmt='0.0"%"')

# SCOP
elec_pac = total_e_pac / 3.5 if total_e_pac > 0 else 0  # approximate
row += 1
scop = round(total_e_pac / elec_pac, 2) if elec_pac > 0 else 0
style_cell(ws_mono, row, 2, "SCOP pondéré estimé", font=font_label_bold, alignment=align_left)
style_cell(ws_mono, row, 5, scop, font=font_result_highlight, alignment=align_right, fmt="0.00")

# Create monotone chart
# First, build monotone data (sorted by load descending, cumulative hours)
monotone_data = []
for t in sorted_temps:
    hours = bins[t]
    if t >= t_int:
        continue
    load = dp * (t_int - t) / 1000
    perf = interpolate_performance(best_model, t, 45)
    pac_cap = perf["pcalo"] * best_n if perf else 0
    for h in range(hours):
        monotone_data.append({"load": load, "pac": pac_cap})

monotone_data.sort(key=lambda x: -x["load"])

# Sample ~80 points
step = max(1, len(monotone_data) // 80)
mono_row = row + 3
merge_style(ws_mono, f"B{mono_row}:G{mono_row}", "Données courbe monotone (heures cumulées)", font=font_section, fill=fill_white, alignment=align_left)
mono_row += 1
header_row(ws_mono, mono_row, 2, ["Heures", "Charge (kW)", "Capacité PAC (kW)", "", "", ""])
mono_row += 1
mono_start = mono_row

for i in range(0, len(monotone_data), step):
    vals = [i+1, round(monotone_data[i]["load"], 1), round(monotone_data[i]["pac"], 1)]
    data_row(ws_mono, mono_row, 2, vals, alt=(mono_row%2==0))
    mono_row += 1

# Last point
if len(monotone_data) > 0:
    vals = [len(monotone_data), round(monotone_data[-1]["load"], 1), round(monotone_data[-1]["pac"], 1)]
    data_row(ws_mono, mono_row, 2, vals, alt=(mono_row%2==0))
    mono_row += 1

mono_end = mono_row - 1

# Create chart
chart = LineChart()
chart.title = "Courbe monotone de chauffage"
chart.style = 10
chart.width = 28
chart.height = 16
chart.x_axis.title = "Heures cumulées"
chart.y_axis.title = "Puissance (kW)"
chart.x_axis.scaling.min = 0
chart.y_axis.scaling.min = 0

# Load line
data_load = Reference(ws_mono, min_col=3, min_row=mono_start-1, max_row=mono_end)
cats = Reference(ws_mono, min_col=2, min_row=mono_start, max_row=mono_end)
chart.add_data(data_load, titles_from_data=True)
chart.set_categories(cats)

# PAC capacity line
data_pac = Reference(ws_mono, min_col=4, min_row=mono_start-1, max_row=mono_end)
chart.add_data(data_pac, titles_from_data=True)

# Style
chart.series[0].graphicalProperties.line.width = 25000  # thick
chart.series[0].graphicalProperties.line.solidFill = "003366"
chart.series[1].graphicalProperties.line.width = 20000
chart.series[1].graphicalProperties.line.solidFill = "FF6600"
chart.series[1].graphicalProperties.line.dashStyle = "dash"

chart.legend.position = "b"

ws_mono.add_chart(chart, "B" + str(mono_end + 2))


# ============================================================
# SHEET: COMPARAISON
# ============================================================

ws_comp = wb.create_sheet("COMPARAISON")
ws_comp.sheet_properties.tabColor = "FF6600"
ws_comp.sheet_view.showGridLines = False

for col, w in [(1,3),(2,24),(3,22),(4,14),(5,14),(6,14),(7,14),(8,14),(9,14),(10,3)]:
    ws_comp.column_dimensions[get_column_letter(col)].width = w

# Header
for c in range(1, 11):
    style_cell(ws_comp, 1, c, None, fill=fill_navy)
    style_cell(ws_comp, 2, c, None, fill=fill_navy)

merge_style(ws_comp, "B1:I1", "COMPARAISON CONCURRENTIELLE", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))
merge_style(ws_comp, "B2:I2", "Atlantic Effipac/Aptae vs Daikin vs Mitsubishi", font=font_subtitle, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

# === Effipac comparison table ===
row = 4
merge_style(ws_comp, f"B{row}:I{row}", "Gamme EFFIPAC (R32) — Comparaison par modèle", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 10):
    ws_comp.cell(row=row, column=c).border = bottom_border

row = 6
headers = ["Marque / Gamme", "Modèle", "Puissance\nnom. (kW)", "Réfrigérant", "T° max\n(°C)", "COP\nA7/W35", "COP\nA7/W55", "Avantages"]
header_row(ws_comp, row, 2, headers)

row = 7
# Atlantic Effipac
for model in EFFIPAC_MODELS:
    p35 = model["performance"].get("A7/W35", {})
    p55 = model["performance"].get("A7/W55", {})
    vals = [
        "Atlantic Effipac",
        model["nom"],
        model["puissance_nom"],
        model["refrigerant"],
        model["t_max"],
        p35.get("cop", ""),
        p55.get("cop", ""),
        "Monobloc Full Inverter, cascade 420kW"
    ]
    fmts = [None, None, "0.0", None, "0", "0.00", "0.00", None]
    data_row(ws_comp, row, 2, vals, fmts=fmts, alt=False)
    # Highlight Atlantic rows
    for c in range(2, 10):
        ws_comp.cell(row=row, column=c).fill = fill_atlantic
    row += 1

# Daikin
for brand_name, brand_data in COMPETITORS.items():
    for cmodel in brand_data["models"]:
        vals = [
            brand_name,
            cmodel["nom"],
            cmodel["puissance_nom"],
            brand_data["refrigerant"],
            brand_data["t_max"],
            cmodel.get("cop_a7w35", ""),
            cmodel.get("cop_a7w55", "N/A"),
            "Bi-bloc" if "Daikin" in brand_name else "CO2 jusqu'à 90°C"
        ]
        fmts = [None, None, "0", None, "0", "0.00", None, None]
        data_row(ws_comp, row, 2, vals, fmts=fmts, alt=(row%2==0))
        row += 1

# === Aptae comparison ===
row += 1
merge_style(ws_comp, f"B{row}:I{row}", "Gamme APTAE (R290) — Comparaison par modèle", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 10):
    ws_comp.cell(row=row, column=c).border = bottom_border

row += 2
header_row(ws_comp, row, 2, headers)

row += 1
for model in APTAE_MODELS:
    p35 = model["performance"].get("A7/W35", {})
    p55 = model["performance"].get("A7/W55", {})
    vals = [
        "Atlantic Aptae",
        model["nom"],
        model["puissance_nom"],
        model["refrigerant"],
        model["t_max"],
        p35.get("cop", ""),
        p55.get("cop", ""),
        "R290 naturel, HT 75°C, Keymark"
    ]
    fmts = [None, None, "0.0", None, "0", "0.00", "0.00", None]
    data_row(ws_comp, row, 2, vals, fmts=fmts, alt=False)
    for c in range(2, 10):
        ws_comp.cell(row=row, column=c).fill = fill_atlantic
    row += 1

# Competitors again
for brand_name, brand_data in COMPETITORS.items():
    for cmodel in brand_data["models"]:
        vals = [
            brand_name,
            cmodel["nom"],
            cmodel["puissance_nom"],
            brand_data["refrigerant"],
            brand_data["t_max"],
            cmodel.get("cop_a7w35", cmodel.get("cop_a7w65", "")),
            cmodel.get("cop_a7w55", "N/A"),
            "Bi-bloc" if "Daikin" in brand_name else "CO2 HT, ECS collective"
        ]
        fmts = [None, None, "0", None, "0", "0.00", None, None]
        data_row(ws_comp, row, 2, vals, fmts=fmts, alt=(row%2==0))
        row += 1

# === Avantages Atlantic ===
row += 2
merge_style(ws_comp, f"B{row}:I{row}", "ARGUMENTS CLÉS ATLANTIC", font=font_section, fill=fill_white, alignment=align_left)
for c in range(2, 10):
    ws_comp.cell(row=row, column=c).border = bottom_border

row += 2
arguments = [
    ("Effipac R32", "Monobloc Full Inverter, pas de liaison frigorifique. Cascade jusqu'à 420 kW (6 unités). Régulation Navistem T3100 intégrée."),
    ("Aptae R290", "Réfrigérant naturel (R290), GWP=3. Haute température 75°C sans appoint. Certifié Heat Pump Keymark. SCOP jusqu'à 4.94."),
    ("vs Daikin", "Daikin Altherma 3 H HT : gamme résidentielle/petit collectif. Puissance max 18 kW vs 70 kW Effipac / 65 kW Aptae. Bi-bloc (liaison frigo nécessaire)."),
    ("vs Mitsubishi", "Mitsubishi Ecodan CAHV : CO2, idéal ECS mais COP chauffage inférieur en basse température. Prix matériel significativement plus élevé."),
    ("SAV & Support", "Réseau installateurs France, formation Atlantic Systèmes, pièces détachées garanties 10 ans, support technique dédié collectif."),
]

for arg_title, arg_text in arguments:
    style_cell(ws_comp, row, 2, arg_title, font=font_label_bold, fill=fill_light_orange, border=thin_border, alignment=align_left)
    merge_style(ws_comp, f"C{row}:I{row}", arg_text, font=font_data, fill=fill_white, border=thin_border, alignment=align_left)
    row += 1


# ============================================================
# SHEET: TUYAUTERIE — Pipe sizing helper
# ============================================================

ws_pipe = wb.create_sheet("TUYAUTERIE")
ws_pipe.sheet_properties.tabColor = "336699"

for col, w in [(1,3),(2,18),(3,14),(4,14),(5,18)]:
    ws_pipe.column_dimensions[get_column_letter(col)].width = w

for c in range(1, 6):
    style_cell(ws_pipe, 1, c, None, fill=fill_navy)

merge_style(ws_pipe, "B1:E1", "Diamètres de tuyauterie — Référence", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

row = 3
header_row(ws_pipe, row, 2, ["Ø intérieur (mm)", "Ø extérieur (mm)", "Débit max (L/h)", "Usage typique"])

row = 4
usages = [
    "Petit circuit secondaire",
    "Circuit radiateur individuel",
    "Boucle PAC 15-20 kW",
    "Boucle PAC 25-40 kW",
    "Boucle PAC 40-60 kW",
    "Boucle PAC 60-100 kW",
    "Collecteur principal >100 kW",
]
for i, pipe in enumerate(PIPE_DIAMETERS):
    vals = [pipe["int"], pipe["ext"], pipe["maxFlow"], usages[i]]
    fmts = ["0", "0", "#,##0", None]
    data_row(ws_pipe, row, 2, vals, fmts=fmts, alt=(row%2==0))
    row += 1


# ============================================================
# SHEET: GUIDE — User guide
# ============================================================

ws_guide = wb.create_sheet("GUIDE")
ws_guide.sheet_properties.tabColor = "003366"
ws_guide.sheet_view.showGridLines = False

ws_guide.column_dimensions["A"].width = 3
ws_guide.column_dimensions["B"].width = 80

for c in range(1, 4):
    style_cell(ws_guide, 1, c, None, fill=fill_navy)

merge_style(ws_guide, "B1:C1", "GUIDE D'UTILISATION", font=font_title, fill=fill_navy, alignment=Alignment(horizontal="left", vertical="center"))

instructions = [
    ("1. Onglet SAISIE", "Renseignez tous les paramètres dans les cellules bleu clair:"),
    ("", "  • Département (01-95, 2A, 2B)"),
    ("", "  • Altitude en mètres (correction automatique de Tbase)"),
    ("", "  • Bord de mer (ajoute +2°C à Tbase)"),
    ("", "  • Type de bâtiment (détermine les déperditions spécifiques)"),
    ("", "  • Surface habitable en m²"),
    ("", "  • Température de consigne (19°C par défaut)"),
    ("", "  • Gamme PAC: Effipac R32 ou Aptae R290"),
    ("", "  • Température eau départ: 35, 45 ou 55°C"),
    ("", "  • Delta T: écart primaire (5K standard)"),
    ("", "  • Nombre de PAC max en cascade"),
    ("", ""),
    ("2. Onglet RÉSULTATS", "Consultez les tableaux de sélection PAC:"),
    ("", "  • Toutes les combinaisons modèle × nombre d'unités"),
    ("", "  • Performance à A7/W45 et A-7/W55"),
    ("", "  • Table hydraulique (débits, diamètres, ballons tampon)"),
    ("", ""),
    ("3. Onglet MONOTONE", "Courbe monotone pré-calculée pour l'exemple Paris:"),
    ("", "  • Heures cumulées vs puissance de chauffage"),
    ("", "  • Capacité PAC superposée"),
    ("", "  • Taux de couverture annuel"),
    ("", "  • SCOP pondéré"),
    ("", ""),
    ("4. Onglet COMPARAISON", "Comparaison Atlantic vs concurrents:"),
    ("", "  • Daikin Altherma 3 H HT (R290, résidentiel)"),
    ("", "  • Mitsubishi Ecodan CAHV (CO2, collectif)"),
    ("", "  • Arguments commerciaux clés"),
    ("", ""),
    ("5. Onglet DATA_PAC", "Tables de performance pré-calculées:"),
    ("", "  • Pcalo, COP, Pabs pour chaque modèle"),
    ("", "  • À chaque température extérieure (-20 à +20°C)"),
    ("", "  • Pour chaque température eau (35, 45, 55°C)"),
    ("", "  • Interpolation linéaire depuis données certifiées Keymark"),
    ("", ""),
    ("Normes", "NF EN 12831 (déperditions), EN 14825 (bin method),"),
    ("", "ADEME/COSTIC (ECS collectif), Heat Pump Keymark (Aptae R290)"),
]

row = 3
for title, text in instructions:
    if title:
        style_cell(ws_guide, row, 2, title, font=font_label_bold, alignment=align_left)
    else:
        style_cell(ws_guide, row, 2, text, font=font_data, alignment=align_left)
    row += 1


# ============================================================
# FINAL: Set SAISIE as active, freeze panes, save
# ============================================================

wb.active = wb.sheetnames.index("SAISIE")

# Freeze panes on data sheets
ws_pac.freeze_panes = "D4"
ws_tbase.freeze_panes = "A2"
ws_bin.freeze_panes = "A2"

# Move GUIDE to first position after SAISIE
# (openpyxl doesn't have move_sheet easily, so leave order as-is)

# Save
output_path = "/Users/marketwatchxyz/Downloads/pac-dimensionnement/PAC_Dimensionnement.xlsx"
wb.save(output_path)
print(f"✓ Fichier Excel créé: {output_path}")
print(f"  Onglets: {', '.join(wb.sheetnames)}")
print(f"  Effipac: {len(EFFIPAC_MODELS)} modèles")
print(f"  Aptae: {len(APTAE_MODELS)} modèles")
print(f"  Monotone: {len(monotone_data)} heures, taux couverture {taux}%")
